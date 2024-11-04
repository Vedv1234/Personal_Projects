"""
Author: Ved Vyas
Co-Author / Exercise provided by / Resources provided by: Mosh Hamedani Python Coding projects (2023)
Functionality of code:
    This is my Excel automation script that helps process pricing data. I created it to automatically
    apply a 10% discount to prices in Excel sheets and visualize the results with a bar chart.
    It's super helpful when I need to handle bulk price updates and want to see the outcome graphically.
"""

# I'm using openpyxl for Excel handling - it's my go-to library for this kind of work
import openpyxl as xl
# I need this for creating those nice bar charts to visualize my data
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    # First, I need to load up the Excel workbook I want to work with
    wb = xl.load_workbook(filename)
    
    # I'm targeting Sheet1 specifically - that's where my data lives
    sheet = wb['Sheet1']
    
    # Now I'll go through each row, starting from row 2 (skipping headers)
    for row in range(2, sheet.max_row + 1):
        # I'm checking if there's actually data in the first column
        # This helps me avoid processing empty rows
        cell = sheet.cell(row, 1)
        if cell.value is None:
            continue
            
        # Here's where I grab the price from column 3
        cell = sheet.cell(row, 3)
        # I apply my 10% discount by multiplying by 0.9
        corrected_price = cell.value * 0.9
        # I store the new price in column 4
        corrected_price_shell = sheet.cell(row, 4)
        corrected_price_shell.value = corrected_price
    
    # Now for the visualization part!
    # I'm selecting all my corrected prices for the chart
    values = Reference(sheet,
              min_row=2, 
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)
              
    # Creating a bar chart to show my price distribution
    chart = BarChart()
    chart.add_data(values)
    # I'm placing the chart at cell E2
    sheet.add_chart(chart, 'e2')
    
    # Finally, saving all my changes back to the file
    wb.save(filename)

def main():
    # I ask the user which file they want to process
    # Note: I add .xlsx extension automatically to make it easier
    filename = input("What is the filename you want to update?") + ".xlsx"
    process_workbook(filename)
    
# This kicks off my whole program
main()